import argparse
from pathlib import Path
from typing import List
from munch import munchify
from PIL import Image
from tqdm import tqdm
import torch
from torchvision.utils import save_image
from torchvision import transforms
from util import set_seed, get_img_list, process_text
from sd3_sampler_TriPS_T import get_solver
from eval import Metric
from torchvision.utils import make_grid
import torchvision.transforms as T
from custom_util import *
from torch.nn import functional as F

# excel save
from datetime import datetime
from openpyxl import Workbook, load_workbook

@torch.no_grad
def precompute(args, prompts:List[str], solver) -> List[torch.Tensor]:
    prompt_emb_set = []
    pooled_emb_set = []
    num_samples = args.num_samples if args.num_samples > 0 else len(prompts)
    for prompt in prompts[:num_samples]:
        # prompt_emb, pooled_emb = solver.encode_prompt(prompt, batch_size=1)
        prompt_emb, pooled_emb = solver.encode_prompt(prompt)
        prompt_emb_set.append(prompt_emb)
        pooled_emb_set.append(pooled_emb)
    return prompt_emb_set, pooled_emb_set

# excel save
def append_metrics_to_xlsx(xlsx_path, row_dict):
    """
    Append one row (row_dict) to an Excel file.
    If file doesn't exist, create it with header.
    """
    xlsx_path = str(xlsx_path)
    headers = list(row_dict.keys())
    values = [row_dict[h] for h in headers]

    try:
        wb = load_workbook(xlsx_path)
        ws = wb.active
        # if empty sheet or no header, write header
        if ws.max_row == 1 and ws.max_column == 1 and ws["A1"].value is None:
            ws.append(headers)
    except FileNotFoundError:
        wb = Workbook()
        ws = wb.active
        ws.title = "metrics"
        ws.append(headers)

    # if header exists but mismatched, you can align here
    ws.append(values)
    wb.save(xlsx_path)

def run(args):
    # load solver
    solver = get_solver(args.method)
    solver.seed = args.seed

    ###### suffix remove ############################
    def sanitize_prompt(s: str, suffix: str) -> str:
        s = s.strip()
        if suffix and s.endswith(suffix):
            s = s[: -len(suffix)].strip()
        return s

    # load text prompts
    prompts = process_text(prompt=args.prompt, prompt_file=args.prompt_file)

    # suffix remove (DIV2K : prompt file // FFHQ : prompt)
    if args.prompt_file is not None:
        prompts = [sanitize_prompt(p, args.prompt_suffix_to_remove) for p in prompts]

    solver.text_enc_1.to('cuda')
    solver.text_enc_2.to('cuda')
    solver.text_enc_3.to('cuda')

    if args.efficient_memory:
        # precompute text embedding and remove encoders from GPU
        # This will allow us 1) fast inference 2) with lower memory requirement (<24GB)
        with torch.no_grad():
            prompt_emb_set, pooled_emb_set = precompute(args, prompts, solver)
            null_emb, null_pooled_emb = solver.encode_prompt([''])

        del solver.text_enc_1
        del solver.text_enc_2
        del solver.text_enc_3
        torch.cuda.empty_cache()
        prompt_embs = [[x, y] for x, y in zip(prompt_emb_set, pooled_emb_set)]
        null_embs = [null_emb, null_pooled_emb]
    else:
        prompt_embs = [[None, None]] * len(prompts)
        null_embs = [None, None]

    print("Prompts are processed.")
    
    solver.vae.to('cuda')
    solver.transformer.to('cuda')

    #-----------------------problem setup------------------------
    device = solver.vae.device
    img_size = args.img_size

    if args.task == 'cs_walshhadamard':
        compress_by = round(1/args.deg_scale)
        from functions.svd_operators import WalshHadamardCS
        A_funcs = WalshHadamardCS(3, img_size, compress_by,
                                    torch.randperm(img_size ** 2, device=device), device)
    elif args.task == 'cs_blockbased':
        cs_ratio = args.deg_scale
        from functions.svd_operators import CS
        A_funcs = CS(3, img_size, cs_ratio, device)
    elif args.task == 'inpainting':
        from functions.svd_operators import Inpainting
        loaded = np.load("inp_masks/FFHQ_mask.npy")
        mask = torch.from_numpy(loaded).to(device).reshape(-1)
        missing_r = torch.nonzero(mask == 0).long().reshape(-1) * 3
        missing_g = missing_r + 1
        missing_b = missing_g + 1
        missing = torch.cat([missing_r, missing_g, missing_b], dim=0)
        A_funcs = Inpainting(3, img_size, missing, device)
    elif args.task == 'inpainting_DIV2K':
        from functions.svd_operators import Inpainting
        loaded = np.load("inp_masks/DIV2k_mask.npy")
        mask = torch.from_numpy(loaded).to(device).reshape(-1)
        missing_r = torch.nonzero(mask == 0).long().reshape(-1) * 3
        missing_g = missing_r + 1
        missing_b = missing_g + 1
        missing = torch.cat([missing_r, missing_g, missing_b], dim=0)
        A_funcs = Inpainting(3, img_size, missing, device)
    elif args.task == 'denoising':
        from functions.svd_operators import Denoising
        A_funcs = Denoising(3, img_size, device)
    elif args.task == 'colorization':
        from functions.svd_operators import Colorization
        A_funcs = Colorization(img_size, device)
    elif args.task == 'sr_averagepooling':
        blur_by = int(args.deg_scale)
        if args.operator_imp == 'SVD':
            from functions.svd_operators import SuperResolution
            A_funcs = SuperResolution(3, img_size, blur_by, device)
        else:
            raise NotImplementedError()

    elif args.task == 'sr_bicubic':
        factor = int(args.deg_scale)
        def bicubic_kernel(x, a=-0.5):
            if abs(x) <= 1:
                return (a + 2) * abs(x) ** 3 - (a + 3) * abs(x) ** 2 + 1
            elif 1 < abs(x) and abs(x) < 2:
                return a * abs(x) ** 3 - 5 * a * abs(x) ** 2 + 8 * a * abs(x) - 4 * a
            else:
                return 0
        k = np.zeros((factor * 4))
        for i in range(factor * 4):
            x = (1 / factor) * (i - np.floor(factor * 4 / 2) + 0.5)
            k[i] = bicubic_kernel(x)
        k = k / np.sum(k)
        kernel = torch.from_numpy(k).float().to(device)
        
        if args.operator_imp == 'SVD':
            from functions.svd_operators import SRConv
            A_funcs = SRConv(kernel / kernel.sum(), 3, img_size, device, stride=factor)
        elif args.operator_imp == 'FFT':                
            from functions.fft_operators import Superres_fft, prepare_cubic_filter
            k = prepare_cubic_filter(1/factor)
            kernel = torch.from_numpy(k).float().to(device)
            A_funcs = Superres_fft(kernel / kernel.sum(), 3, img_size, device, stride=factor)
        else:
            raise NotImplementedError()

    elif args.task == 'deblur_uni':
        if args.operator_imp == 'SVD':
            from functions.svd_operators import Deblurring
            A_funcs = Deblurring(torch.Tensor([1 / 9] * 9).to(device), 3, img_size, device)
        elif args.operator_imp == 'FFT':
            from functions.fft_operators import Deblurring_fft
            A_funcs = Deblurring_fft(torch.Tensor([1 / 9] * 9).to(device), 3, img_size, device)
        else:
            raise NotImplementedError()

    elif args.task == 'deblur_gauss':
        sigma = args.deg_scale
        pdf = lambda x: torch.exp(torch.Tensor([-0.5 * (x / sigma) ** 2]))
        size = 61
        ker = []
        for k in range(-size//2, size//2):
            ker.append(pdf(k))
        kernel = torch.Tensor(ker).to(device)

        if args.operator_imp == 'SVD':
            from functions.svd_operators import Deblurring
            A_funcs = Deblurring(kernel / kernel.sum(), 3, img_size, device)
        elif args.operator_imp == 'FFT':
            from functions.fft_operators import Deblurring_fft
            A_funcs = Deblurring_fft(kernel / kernel.sum(), 3, img_size, device)
        else:
            raise NotImplementedError()

    elif args.task == 'deblur_aniso':
        sigma = 20
        pdf = lambda x: torch.exp(torch.Tensor([-0.5 * (x / sigma) ** 2]))
        kernel2 = torch.Tensor([pdf(-4), pdf(-3), pdf(-2), pdf(-1), pdf(0), pdf(1), pdf(2), pdf(3), pdf(4)]).to(device)
        sigma = 1
        pdf = lambda x: torch.exp(torch.Tensor([-0.5 * (x / sigma) ** 2]))
        kernel1 = torch.Tensor([pdf(-4), pdf(-3), pdf(-2), pdf(-1), pdf(0), pdf(1), pdf(2), pdf(3), pdf(4)]).to(device)

        if args.operator_imp == 'SVD':
            from functions.svd_operators import Deblurring2D
            A_funcs = Deblurring2D(kernel1 / kernel1.sum(), kernel2 / kernel2.sum(), 3, img_size, device)
        elif args.operator_imp == 'FFT':
            # unlike when using 'SVD' mode, here you can implement any 2D kernel that you want (not just seperable kernels)
            from functions.fft_operators import Deblurring_fft
            kernel = torch.matmul(kernel1[:,None],kernel2[None,:])
            A_funcs = Deblurring_fft(kernel / kernel.sum(), 3, img_size, device)
        else:
            raise NotImplementedError()

    elif args.task == 'deblur_motion':
        from functions.motionblur.motionblur import Kernel
        if args.operator_imp == 'FFT':
            from functions.fft_operators import Deblurring_fft
        else:
            raise ValueError("set operator_imp = FFT")

    else:
        raise ValueError("degradation type not supported")

    # solve problem
    tf = transforms.Compose([
        transforms.Resize(args.img_size),
        transforms.CenterCrop(args.img_size),
        transforms.ToTensor()
        ])

    pbar = tqdm(get_img_list(args.img_path), desc="Solving")
    for i, path in enumerate(pbar):

        img_name = path.stem
        img_num = int(img_name)
        prompt_idx = img_num - 1

        img = tf(Image.open(path).convert('RGB'))
        img = img.unsqueeze(0).to(solver.vae.device)
        img = img * 2 - 1
        
        if args.task == 'deblur_motion':
            from functions.motionblur.motionblur import Kernel
            if args.operator_imp == 'FFT':
                from functions.fft_operators import Deblurring_fft
            else:
                raise ValueError("set operator_imp = FFT")
            np.random.seed(seed=i * 10)  # for reproducibility of blur kernel for each image
            kernel = torch.from_numpy(Kernel(size=(args.deg_scale, args.deg_scale), intensity=0.5).kernelMatrix)
            A_funcs = Deblurring_fft(kernel / kernel.sum(), 3, args.img_size, solver.transformer.device)
        
        y = A_funcs.A(img)
        y = y + args.noise_std * torch.randn(y.shape, device=y.device, generator=torch.Generator(y.device).manual_seed(args.seed))
        
        out, images = solver.sample(measurement=y,
                                          operator=A_funcs,
                                          task=args.task,
                                          prompts=prompts[i] if len(prompts)>1 else prompts[0],
                                          NFE=args.NFE,
                                          img_shape=(args.img_size, args.img_size),
                                          cfg_scale=args.cfg_scale,
                                          prompt_embs=prompt_embs[i] if len(prompt_embs)>1 else prompt_embs[0],
                                          null_embs=null_embs,
                                          step_scale=args.step_scale,
                                          step_scale_2=args.step_scale_2,
                                          inner_steps=args.inner_steps,
                                          sigma_y=args.noise_std,
                                          stochasticity_weight=args.stochasticity_weight,
                                          workdir=args.workdir,
                                          function_dc=args.function_dc,
                                          function_cfg=args.function_cfg,
                                          function_sto=args.function_sto,
                                          img_num=img_num
        )

        # Save results
        i = img_num

        if args.task in ['sr_bicubic', 'inpainting', 'inpainting_DIV2K', 'cs_walshhadamard', 'cs_blockbased', 'deblur_gauss']: # modify
            save_image(A_funcs.At(y).reshape(img.shape),
                    args.workdir.joinpath(f'input1/{str(i).zfill(4)}.png'),
                    normalize=True)
        else:
            save_image(y.reshape(img.shape),
                    args.workdir.joinpath(f'input1/{str(i).zfill(4)}.png'),
                    normalize=True)

        save_image(out,
                args.workdir.joinpath(f'recon/{str(i).zfill(4)}.png'),
                normalize=True)
        save_image(img,
                args.workdir.joinpath(f'label/{str(i).zfill(4)}.png'),
                normalize=True)

        # Store processes
        fname = str(i).zfill(5) + f'.png'
        images = torch.cat(images, dim=0)
        grid = make_grid(images, nrow=5, normalize=True)
        to_pil = T.ToPILImage()
        grid_img = to_pil(grid)
        grid_img = grid_img.resize((grid_img.width // 2, grid_img.height // 2))
        grid_img.save(args.workdir.joinpath('process', fname))

    # ============================
    # Evaluation
    # ============================
    metric = Metric(['psnr', 'lpips_FLAIR', 'lpips_FlowDPS', 'ssim'])
    recon_path_1 = args.workdir.joinpath('recon')
    label_path = args.workdir.joinpath('label')

    psnr_val_1, lpips_val_1, lpips_val_2, ssim_val_1 = metric(recon_path_1, label_path)

    print(f"\n==== 1st stage Evaluation Results ====")
    print(f"PSNR : {psnr_val_1:.4f} dB")
    print(f"LPIPS-FLAIR: {lpips_val_1:.4f}")
    print(f"LPIPS-FlowDPS: {lpips_val_2:.4f}")
    print(f"SSIM : {ssim_val_1:.4f}")

    result_file = args.workdir.joinpath("eval_results_1st_stage.txt")
    with open(result_file, "a") as f:
        f.write(f"\n==== {i} sample 1st stage ====\n")
        f.write(f"PSNR : {psnr_val_1:.4f} dB\n")
        f.write(f"LPIPS-FLAIR: {lpips_val_1:.4f}\n")
        f.write(f"LPIPS-FlowDPS: {lpips_val_2:.4f}\n")
        f.write(f"SSIM : {ssim_val_1:.4f}\n")

    ###### key point #####################################
    eval_score = (psnr_val_1 - 20.0)/20 - lpips_val_2

    # ============================
    # Append summary metrics (one row per experiment folder)
    # ============================
    base_dir = args.workdir.parent  # e.g., .../motion_deblur.../
    summary_xlsx = base_dir / "summary_metrics.xlsx"

    row = {
        "time": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "task": args.task,
        "img_path": str(args.img_path),
        "workdir": str(args.workdir),
        "seed": args.seed,
        "NFE": args.NFE,
        "deg_scale": args.deg_scale,
        "noise_std": args.noise_std,
        "step_scale": args.step_scale,
        "step_scale_2": args.step_scale_2,
        "inner_steps": args.inner_steps,
        "function_dc": args.function_dc,
        "function_cfg": args.function_cfg,
        "function_sto": args.function_sto,
        "PSNR": float(psnr_val_1),
        "LPIPS-FLAIR": float(lpips_val_1),
        "LPIPS-FlowDPS": float(lpips_val_2),
        "SSIM": float(ssim_val_1),
        "eval_score": float(eval_score),
    }
    append_metrics_to_xlsx(summary_xlsx, row)
    print(f"[Summary] appended to: {summary_xlsx}")

if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    # sampling params
    parser.add_argument('--seed', type=int, default=0)
    parser.add_argument('--NFE', type=int, default=28)
    parser.add_argument('--cfg_scale', type=float, default=2.0)
    parser.add_argument('--img_size', type=int, default=768)
    # workdir params
    parser.add_argument('--workdir', type=Path, default='workdir_demo_TriPS_T')
    # data params
    parser.add_argument('--img_path', type=Path)
    parser.add_argument('--prompt', type=str, default=None)
    parser.add_argument('--prompt_file', type=str, default=None)
    parser.add_argument('--num_samples', type=int, default=-1)
    # problem params
    parser.add_argument('--task', type=str, default='sr_avgpool')
    parser.add_argument('--method', type=str, default='TriPS_T')
    parser.add_argument('--deg_scale', type=int, default=12)
    parser.add_argument('--noise_std', type=float, default=0.03)
    # solver params
    parser.add_argument('--efficient_memory',default=False, action='store_true')
    parser.add_argument('--attn_enforce', type=float, default=1.3)
    parser.add_argument('--step_scale', type=float, default=20.0)
    parser.add_argument('--step_scale_2', type=float, default=10.0)
    parser.add_argument('--inner_steps', type=int, default=6)
    parser.add_argument('--stochasticity_weight', type=float, default=1.0)
    # Added for operator
    parser.add_argument(
        "--operator_imp", type=str, default="FFT", help="SVD | FFT"  # TODO: add CG support
    )
    parser.add_argument('--prompt_suffix_to_remove', type=str, default=', high-resolution, 8k')

    # Scheduler Template functions
    parser.add_argument('--function_dc', type=str, default='linear')
    parser.add_argument('--function_cfg', type=str, default='linear')
    parser.add_argument('--function_sto', type=str, default='linear')

    args = parser.parse_args()

    # workdir creation and seed setup
    set_seed(args.seed)
    args.workdir.joinpath('input1').mkdir(parents=True, exist_ok=True)
    args.workdir.joinpath('recon').mkdir(parents=True, exist_ok=True)
    args.workdir.joinpath('label').mkdir(parents=True, exist_ok=True)
    args.workdir.joinpath('process').mkdir(parents=True, exist_ok=True)
    # run main script
    run(args)






